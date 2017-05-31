import re
import random
import openpyxl

class Manager(object):

	def __init__(self,model_file):

		self.__getElement = dict()
		self.__updateList = list()
		self.__initial = dict()

		wb = openpyxl.load_workbook(model_file)
		ws = wb.active

		max_state = 3
		if ws.cell(row=1, column=4).value != None:
			max_state = ws.cell(row=1, column=4).value

		curr_row = 2
		while ws.cell(row=curr_row, column=1).value != None:

			# print('Parsing',ws.cell(row=curr_row, column=1).value)

			val = 1
			if ws.cell(row=curr_row,column=6).value != None:
				val = ws.cell(row=curr_row,column=6).value

			X = ws.cell(row=curr_row,column=1).value
			A = ws.cell(row=curr_row,column=2).value
			I = ws.cell(row=curr_row,column=3).value
			
			X = '' if X==None else X.strip()
			A = '' if A==None else A.strip()
			I = '' if I==None else I.strip()

			ele = gateNode(X,A,I,val,max_state)
			self.__initial[X] = val
			self.__getElement[X] = ele
			if A!='' or I!='':
				self.__updateList += [ele]

			curr_row += 1

	def set_initial(self):

		for name in self.__getElement:
			self.__getElement[name].set_value(self.__initial[name])

	def run_simulation(self,simtype,m,simStep,outName,**kwargs):
		
		outMode = kwargs["outMode"] if "outMode" in kwargs else 1

		output_file = open(outName,'w')
		freq_sum = dict()
		for key in self.__getElement:
			freq_sum[key] = (simStep+1) * [0]
			freq_sum[key][0] = self.__getElement[key].get_value() * runs


		for run in range(runs):
			if outMode!=3:
				output_file.write('Run #'+str(run)+'\n')
				
			self.set_initial()
			memo = dict()
			for key in self.__getElement:
				memo[key] = [self.__getElement[key].get_value()]

			for step in range(1,simStep+1):
				self.ra_update()
				for key in self.__getElement:
					memo[key] += [self.__getElement[key].get_value()]
					freq_sum[key][step] += self.__getElement[key].get_value()

			if outMode!=3: 
				for name in sorted(self.__getElement):
					output_file.write(name+' '+' '.join([str(x) for x in memo[name]])+'\n')

		output_file.write('\nFrequency Summary:\n')
		for name in sorted(self.__getElement):
			output_file.write(name+' '+' '.join([str(x) for x in freq_sum[name]])+'\n')


	def run_simulation_checker(self,simtype,simStep,outName):

		output_file = open(outName,'w')
		output_file.write('# time ')
		for key in sorted(self.__getElement):
			output_file.write(key+'_0 '+key+'_1 ')
		output_file.write('step\n')

		self.print_value(output_file,0)
		for step in range(1,simStep+1):
			self.ra_update()
			self.print_value(output_file,step)

		output_file.close()		


	def ra_update(self):
		
		update_ele = random.choice(self.__updateList)
		# print('update: ',update_ele.get_name(),update_ele.get_value())
		update_ele.update(self.__getElement)

	def print_value(self,output_file,step):

		output_file.write(str(step)+'  ')
		for key in sorted(self.__getElement):
			val = self.__getElement[key].get_value()
			output_file.write(str(val&1)+' ')
			output_file.write(str((val&2)>>1)+' ')
		output_file.write(str(step)+'\n')



class gateNode(object):

	def __init__(self,X,A,I,curr_val,max_state=3):
		self.__regulated = X.strip()
		self.__act = re.sub('\s','',A)
		self.__inh = re.sub('\s','',I)
		self.__name_list = self.create_name_list(X.strip(),A.strip(),I.strip())
		self.__name_to_value = dict()
		self.__curr_val = curr_val
		self.__max_state = max_state

	##### Get functions #####

	def get_name(self):
		return self.__regulated

	def get_max_state(self):
		return self.__max_state

	def get_name_list(self):
		return self.__name_list

	def get_value(self):
		return self.__curr_val

	#########################

	def set_value(self,val):
		self.__curr_val = val

	def create_name_list(self,X,A,I):
		names = set([X])
		act_set = set(re.findall(r'[\w_@]+',A))
		inh_set = set(re.findall(r'[\w_@]+',I))
		return sorted(list(act_set-names)) + sorted(list(inh_set-act_set-names)) + list(names)

	def update(self,getElement):
		self.__name_to_value.clear()
		for name in self.__name_list:
			self.__name_to_value[name] = getElement[name].get_value()
			# print(name,self.__name_to_value[name])
		# print(self.evaluate())
		self.__curr_val = self.evaluate()

	def evaluate(self):
		y_act = self.eval_act(self.__act,0)
		y_inh = self.eval_inh(self.__inh)
		
		if y_act==0 and y_inh==0:
			gradient = 0
		elif y_act>y_inh:
			gradient = 1
		else:
			gradient = -1
		return sorted([0, self.__name_to_value[self.__regulated]+gradient, self.__max_state-1])[1]

	def eval_act(self,act_rule,layer):
		act_list = self.split_comma_outside_parentheses(act_rule)
		y_init = list()
		y_sum = list()

		for act_element in act_list:
			if act_element[0]=='{' and act_element[-1]=='}':
				assert(layer==0)
				y_init = self.eval_act(act_element[1:-1],1)
			elif act_element[0]=='{' and act_element[-1]==']':
				parentheses = 0
				cut_point = 0
				for index in range(len(act_element)): # Find the cut point between {} and []
					if act_element[index]=='{':
						parentheses += 1
					elif act_element[index]=='}':
						parentheses -= 1
					if parentheses==0:
						cut_point = index
						break
				y_must = self.eval_act(act_element[1:cut_point],1)
				y_enhance = self.eval_act(act_element[cut_point+2:-1],1)
				y_sum += [0 if all([y==0 for y in y_must])==True \
					else sorted([0, max(y_must)+max(y_enhance), self.__max_state-1])[1]]
			elif act_element[0]=='(' and act_element[-1]==')':
				y_and = [x \
					for and_entity in self.split_comma_outside_parentheses(act_element[1:-1]) \
					for x in self.eval_act(and_entity,1)]
				y_sum += [min(y_and)]
			else:# Single Elements
				assert(act_element.find(',')==-1)
				if act_element[-1]=='+':
					if act_element[0]=='!':
						y_sum += [int(not self.__name_to_value[act_element[1:-1]]==self.__max_state-1)]
					else:
						y_sum += [int(self.__name_to_value[act_element[:-1]]==self.__max_state-1)*2]
				elif act_element[0]=='!':
					y_sum += [int(not bool(self.__name_to_value[act_element[1:]]))]
				else:
					y_sum += [self.__name_to_value[act_element]]

		#print(act_rule + ': ' + str(y_sum))
		if layer==0:
			if self.__name_to_value[self.__regulated]==0 and len(y_init)!=0 and all([y==0 for y in y_init])==True:
				return 0
			else:
				return sum(y_init) + sum(y_sum)
		else:
			return y_sum

	def eval_inh(self,inh_rule):
		inh_list = self.split_comma_outside_parentheses(inh_rule)
		y_sum = 0

		for inh_element in inh_list:
			if inh_element[0]=='(' and inh_element[-1]==')':
				y_and = list()
				for and_entity in self.split_comma_outside_parentheses(inh_element[1:-1]):
					res = self.eval_inh(and_entity)
					if type(res) is int:
						y_and += [res]
					else:
						y_and += [self.eval_inh(and_entity)]
				y_sum += min(y_and)
			else:# Single Elements
				assert(inh_element.find(',')==-1)
				if inh_element[-1]=='+':
					if inh_element[0]=='!':
						y_sum += int(not self.__name_to_value[inh_element[1:-1]]==self.__max_state-1)
					else:
						y_sum += int(self.__name_to_value[inh_element[:-1]]==self.__max_state-1)*2
				elif inh_element[0]=='!':
					y_sum += int(not bool(self.__name_to_value[inh_element[1:]]))
				else:
					y_sum += self.__name_to_value[inh_element]

		return y_sum


	def split_comma_outside_parentheses(self,sentence):
		final_list = list()
		parentheses = 0
		start = 0
		for index in range(len(sentence)):
			char = sentence[index]
			if index==len(sentence)-1:
				final_list.append(sentence[start:index+1])
			elif char=='(' or char=='{' or char=='[':
				parentheses += 1
			elif char==')' or char=='}' or char==']':
				parentheses -= 1
			elif (char==',' and parentheses==0):
				final_list.append(sentence[start:index])
				start = index+1
		return final_list
